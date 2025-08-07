import sys
import os
import json
import time
import re
import pandas as pd
import requests
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QLabel, QFileDialog,
    QMessageBox, QTextEdit, QInputDialog, QTabWidget, QMainWindow, QProgressBar, QCheckBox, QLineEdit
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSize
from PyQt5.QtGui import QFont, QTextCursor
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font as ExcelFont
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QMessageBox, QTableView, QMessageBox, QHBoxLayout
from datetime import datetime, timedelta, timezone
from PyQt5.QtGui import QStandardItemModel, QStandardItem, QIcon, QFont
from PyQt5.QtCore import QSortFilterProxyModel
from dotenv import load_dotenv

load_dotenv()  # .env 있으면 로드

VT_URL = 'https://www.virustotal.com/api/v3/ip_addresses/'

def load_api_keys_from_env():
    def split_keys(env_name):
        val = os.getenv(env_name, "")
        return [k.strip() for k in val.split(",") if k.strip()]

    return {
        "AIPS": split_keys("VT_KEYS_AIPS"),
        "HIPS": split_keys("VT_KEYS_HIPS"),
        "웹방화벽": split_keys("VT_KEYS_WAF")
    }

API_KEYS = load_api_keys_from_env()


CACHE_FILE = 'vt_cache.json'
CACHE_TTL = timedelta(hours=24)

def load_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_cache(cache):
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)



class AnalysisThread(QThread):
    finished = pyqtSignal(str, str)  # stats_msg, output_file
    error = pyqtSignal(str)
    progress = pyqtSignal(int, str)  # percent, current_ip

    def __init__(self, path, mode, parent=None):
        super().__init__(parent)
        self.path = path
        self.mode = mode

    def run(self):
        try:
            if self.mode == "AIPS":
                df, count = self.process_ips(self.path)
                stats_msg = f"🛡️ 차단된 IP(V) 건수: {count}건"
            elif self.mode == "HIPS":
                df, count = self.process_ips(self.path)
                stats_msg = f"🛡️ 차단된 IP(V) 건수: {count}건"
            else:
                df, count = self.process_waf(self.path)
                stats_msg = f"🧹 제거된 대응 패턴 건수: {count}건"

            output_file = self.path.replace('.csv', '_result.xlsx')
            df.to_excel(output_file, index=False)

            wb = load_workbook(output_file)
            ws = wb.active
            for column_cells in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                col_letter = get_column_letter(column_cells[0].column)
                ws.column_dimensions[col_letter].width = max_length + 2
            wb.save(output_file)

            self.finished.emit(stats_msg, output_file)
        except Exception as e:
            self.error.emit(str(e))

    def query_virustotal(self, ip_list, mode):
        DELAY_BY_MODE = {"AIPS": 15, "HIPS": 15, "웹방화벽": 15}
        api_key_list = API_KEYS.get(mode)
        if not api_key_list:
            raise ValueError(f"알 수 없는 모드: {mode}")

        # 1) 중복 제거
        unique_ips = list(dict.fromkeys(ip_list))

        # 2) 캐시 로드 & TTL 검사
        cache = load_cache()
        now = datetime.now(timezone.utc)
        to_query = []
        for ip in unique_ips:
            entry = cache.get(ip)
            if not entry:
                to_query.append(ip)
            else:
                entry_ts = datetime.fromisoformat(entry["ts"])
                # tz-naive 값은 UTC로 간주
                if entry_ts.tzinfo is None:
                    entry_ts = entry_ts.replace(tzinfo=timezone.utc)
                if now - entry_ts >= CACHE_TTL:
                    to_query.append(ip)

        # 3) 실제 API 호출은 to_query만
        total = len(to_query)
        self.total_count = total
        self.current_index = 0

        for idx, ip in enumerate(to_query):
            self.current_index = idx
            percent = int((idx + 1) / total * 100) if total else 100
            self.progress.emit(percent, ip)

            attempt = 0
            entry_data = None

            while attempt < 3:
                api_key = api_key_list[idx % len(api_key_list)]
                try:
                    response = requests.get(VT_URL + ip,
                                            headers={"x-apikey": api_key},
                                            timeout=16)
                    if response.status_code == 200:
                        print(f"[VT] {ip} 조회 성공")
                        # 정상 응답일 때만 JSON 파싱
                        data = response.json().get("data", {}).get("attributes", {})
                        stats = data.get("last_analysis_stats", {})
                        entry_data = {
                            "Malicious": stats.get("malicious", 0),
                            "Suspicious": stats.get("suspicious", 0),
                            "Phishing": stats.get("phishing", 0),
                            "Clean": stats.get("clean", 0),
                            "Harmless": stats.get("harmless", 0),
                            "Undetected": stats.get("undetected", 0),
                            "Country": data.get("country", "N/A"),
                            "ASN": data.get("asn", "N/A"),
                            "AS_Owner": data.get("as_owner", "N/A")
                        }
                        break  # 성공했으니 retry 루프 종료
                    else:
                        print(f"[VT] {ip} 조회 실패 (코드 {response.status_code})")
                        entry_data = {k: "Error" for k in [
                            "Malicious","Suspicious","Phishing","Clean",
                            "Harmless","Undetected","Country","ASN","AS_Owner"
                        ]}
                        break  # HTTP 에러 코드면 retry 하지 않고 빠져나감
                except Exception as e:
                    print(f"[VT] {ip} 조회 오류: {e}")
                    attempt += 1
                    time.sleep(2 ** attempt)

            if entry_data is None:
                entry_data = {k: "Error" for k in [
                    "Malicious", "Suspicious", "Phishing", "Clean",
                    "Harmless", "Undetected", "Country", "ASN", "AS_Owner"
                ]}

            # 4) 캐시에 저장
            cache[ip] = {"ts": now.isoformat(), "data": entry_data}

            time.sleep(DELAY_BY_MODE.get(mode, 15))

        # 5) 캐시 영구 저장 & 최종 반환
        save_cache(cache)
        return {ip: cache[ip]["data"] for ip in unique_ips}




    def process_ips(self, path):
        df = pd.read_csv(path, encoding='utf-8', sep='\t')
        df = df.fillna('').astype(str)
        df = df.apply(lambda col: col.str.strip().str.replace('\n', '', regex=False))
        df['공격자_IP'] = df['공격자'].apply(lambda x: re.search(r'(\d{1,3}(?:\.\d{1,3}){3})', x).group(1) if re.search(r'(\d{1,3}(?:\.\d{1,3}){3})', x) else '')
        df = df[~df['공격자_IP'].str.startswith(('10.20.', '10.30.', '10.40.', '192.168.10.', '10.1.'))].copy()
        df['차단'] = df['차단'].str.strip().str.upper()
        blocked_count = df[df['차단'] == 'V'].shape[0]
        df = df[df['차단'] != 'V']
        df['공격자_공격명'] = df['공격자_IP'] + ' || ' + df['부가정보']
        df = df.drop_duplicates(subset='공격자_공격명')
        df_final = df['공격자_공격명'].str.split(r' \|\| ', expand=True)
        df_final.columns = ['IP', 'Attack_Type']
        df_final['Attack_Type'] = df_final['Attack_Type'].str.replace(r'[\[\]]', '', regex=True).str.strip()
        df_grouped = df_final.groupby('IP')['Attack_Type'].apply(lambda x: ', '.join(sorted(set(x)))).reset_index()
        ip_info_cache = self.query_virustotal(df_grouped['IP'].tolist(), mode=self.mode)
        enriched = []
        for _, row in df_grouped.iterrows():
            ip = row['IP']
            enriched.append({
                "IP": ip,
                "Attack_Type": row['Attack_Type'],
                **ip_info_cache.get(ip, {})
            })
        return pd.DataFrame(enriched), blocked_count

    def process_waf(self, path):
        df = pd.read_csv(path, encoding='utf-8-sig', sep=',')
        if df.columns[0] == '':
            df = df.drop(columns=df.columns[0])
        df.columns = df.columns.str.strip()
        df['출발지 주소'] = df['출발지 주소'].str.split(':').str[0]
        df = df[df['출발지 주소'].apply(lambda x: bool(re.match(r'^(\d{1,3}\.){3}\d{1,3}$', str(x))))]
        df['출발지_IP'] = df['출발지 주소'].apply(lambda x: re.search(r'(\d{1,3}(?:\.\d{1,3}){3})', str(x)).group(1) if re.search(r'(\d{1,3}(?:\.\d{1,3}){3})', str(x)) else '')
        df = df[~df['출발지_IP'].str.startswith(('10.20.', '10.30.', '10.40.', '192.168.10.', '10.1.'))]
        df = df[['룰', '출발지 주소', '대응']]
        patterns_to_remove = ['연결 끊기', '오류 코드', '페이지 리다이렉션', '사용자 페이지']
        removed_count = df['대응'].str.contains('|'.join(patterns_to_remove), na=False).sum()
        df = df[~df['대응'].str.contains('|'.join(patterns_to_remove), na=False)]
        df = df.drop(columns=['대응'])

        df = df[['룰', '출발지 주소']].drop_duplicates().copy()
        
        grouped = (
        df.groupby('출발지 주소')['룰']
        .apply(lambda s: ', '.join(sorted(set(s.astype(str).str.strip()))))
        .reset_index()
    )

        # VirusTotal은 유니크 IP로 한 번만 조회
        ip_info_cache = self.query_virustotal(grouped['출발지 주소'].tolist(), mode="웹방화벽")

        # 조회 결과 병합
        results = []
        for _, row in grouped.iterrows():
            ip = row['출발지 주소']
            rules_joined = row['룰']  # "SQLi, XSS, ..." 형태
            info = ip_info_cache.get(ip, {})
            results.append({
                "출발지 주소": ip,
                "룰": rules_joined,
                **info
            })

        return pd.DataFrame(results), removed_count

class HistoryTab(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.setLayout(QVBoxLayout())
        self.layout().setContentsMargins(10, 10, 10, 10)

        # ── 버튼 바 ──
        btn_bar = QWidget()
        btn_layout = QHBoxLayout(btn_bar)
        btn_layout.setSpacing(10)
        btn_layout.setContentsMargins(0,0,0,0)

        self.btn_export = QPushButton(QIcon("icons/export.png"), "내보내기")
        self.btn_clear  = QPushButton(QIcon("icons/clear.png"),  "전체 지우기")
        for btn in (self.btn_export, self.btn_clear):
            btn.setIconSize(QSize(20,20))
            btn.setCursor(Qt.PointingHandCursor)
            btn.setMinimumHeight(30)
            btn_layout.addWidget(btn)

        btn_layout.addStretch()
        btn_layout.addWidget(QLabel("검색:"))
        self.search = QLineEdit()
        self.search.setPlaceholderText("내용 검색")
        self.search.setFixedWidth(200)
        btn_layout.addWidget(self.search)

        self.layout().addWidget(btn_bar)

        # ── 테이블 뷰 ──
        self.model = QStandardItemModel(0, 2, self)
        self.model.setHorizontalHeaderLabels(["시간", "로그 메시지"])
        self.proxy = QSortFilterProxyModel(self)
        self.proxy.setSourceModel(self.model)
        self.proxy.setFilterKeyColumn(1)  # 메시지 컬럼만 검색

        self.table = QTableView()
        self.table.setModel(self.proxy)
        self.table.setSelectionBehavior(QTableView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().hide()
        self.table.setEditTriggers(QTableView.NoEditTriggers)
        self.table.setFont(QFont("Segoe UI", 10))
        self.layout().addWidget(self.table, 1)

        # ── 시그널 연결 ──
        self.btn_export.clicked.connect(self.export_log)
        self.btn_clear.clicked.connect(self.clear_log)
        self.search.textChanged.connect(self.proxy.setFilterFixedString)

        self.load_history()
        self.apply_styles()

    def apply_styles(self):
        self.setStyleSheet("""
            QPushButton {
                background-color: #1976d2; color: white; border: none;
                padding: 6px 12px; border-radius: 4px;
            }
            QPushButton:hover { background-color: #1565c0; }
            QLineEdit {
                padding: 4px; border: 1px solid #ccc; border-radius: 4px;
            }
            QTableView {
                background-color: #fff; gridline-color: #e0e0e0;
                selection-background-color: #1976d2; selection-color: #fff;
            }
            QHeaderView::section {
                background-color: #f5f5f5; padding: 6px; border: none;
                font-weight: bold;
            }
        """)

    def load_history(self):
        self.model.removeRows(0, self.model.rowCount())
        if not os.path.exists("analysis_history.log"):
            return
        with open("analysis_history.log", "r", encoding="utf-8") as f:
            for line in f:
                ts, msg = line.strip().split(" ", 1)
                time_item = QStandardItem(ts)
                msg_item  = QStandardItem(msg)
                time_item.setTextAlignment(Qt.AlignCenter)
                self.model.appendRow([time_item, msg_item])

    def export_log(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "로그 내보내기", "analysis_history_export.log", "Log Files (*.log *.txt)"
        )
        if not path:
            return
        with open("analysis_history.log", "r", encoding="utf-8") as src, \
             open(path, "w", encoding="utf-8") as dst:
            dst.write(src.read())
        QMessageBox.information(self, "내보내기 완료", f"로그가 저장되었습니다:\n{path}")

    def clear_log(self):
        reply = QMessageBox.question(
            self, "전체 지우기", "모든 로그를 지우시겠습니까?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        if os.path.exists("analysis_history.log"):
            os.remove("analysis_history.log")
        self.load_history()


class CacheTab(QWidget):
    def __init__(self):
        super().__init__()
        self.setLayout(QVBoxLayout())
        self.layout().setContentsMargins(10, 10, 10, 10)

        # ── 버튼 바 ──
        button_bar = QWidget()
        btn_layout = QHBoxLayout(button_bar)
        btn_layout.setSpacing(10)
        btn_layout.setContentsMargins(0, 0, 0, 0)

        self.btn_refresh = QPushButton(QIcon("icons/refresh.png"), "새로고침")
        self.btn_delete  = QPushButton(QIcon("icons/delete.png"),  "선택 삭제")
        self.btn_clear   = QPushButton(QIcon("icons/clear.png"),   "전체 초기화")
        for btn in (self.btn_refresh, self.btn_delete, self.btn_clear):
            btn.setIconSize(QSize(20,20))
            btn.setCursor(Qt.PointingHandCursor)
            btn.setMinimumHeight(30)
            btn_layout.addWidget(btn)

        btn_layout.addStretch()

        self.search = QLineEdit()
        self.search.setPlaceholderText("IP 또는 Timestamp 검색")
        self.search.setFixedWidth(200)
        btn_layout.addWidget(QLabel("검색:"))
        btn_layout.addWidget(self.search)

        self.layout().addWidget(button_bar)

        # ── 테이블 뷰 ──
        self.model = QStandardItemModel(0, 2, self)
        self.model.setHorizontalHeaderLabels(["IP 주소", "DATETIME"])
        self.proxy = QSortFilterProxyModel(self)
        self.proxy.setSourceModel(self.model)
        self.proxy.setFilterKeyColumn(-1)

        self.table = QTableView()
        self.table.setModel(self.proxy)
        self.table.setSelectionBehavior(QTableView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().hide()
        self.table.setEditTriggers(QTableView.NoEditTriggers)
        self.table.setFont(QFont("Segoe UI", 10))
        self.layout().addWidget(self.table, 1)

        # ── 시그널 연결 ──
        self.btn_refresh.clicked.connect(self.load_cache)
        self.btn_delete.clicked.connect(self.delete_selected)
        self.btn_clear.clicked.connect(self.clear_all)
        self.search.textChanged.connect(self.proxy.setFilterFixedString)

        self.load_cache()
        self.apply_styles()

    def apply_styles(self):
        self.setStyleSheet("""
            QPushButton {
                background-color: #1976d2; color: white; border: none;
                padding: 6px 12px; border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1565c0;
            }
            QLineEdit {
                padding: 4px; border: 1px solid #ccc; border-radius: 4px;
            }
            QTableView {
                background-color: #fff; gridline-color: #e0e0e0;
                selection-background-color: #1976d2; selection-color: #fff;
            }
            QHeaderView::section {
                background-color: #f5f5f5; padding: 6px; border: none;
                font-weight: bold;
            }
        """)

    def load_cache(self):
        self.model.removeRows(0, self.model.rowCount())
        cache = {}
        if os.path.exists(CACHE_FILE):
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                cache = json.load(f)
        for ip, entry in cache.items():
            row = [QStandardItem(ip), QStandardItem(entry.get("ts", ""))]
            for item in row:
                item.setTextAlignment(Qt.AlignCenter)
            self.model.appendRow(row)

    def delete_selected(self):
        sel = self.table.selectionModel().selectedRows()
        if not sel:
            return
        ips = [self.proxy.data(idx) for idx in sel]
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            cache = json.load(f)
        for ip in ips:
            cache.pop(ip, None)
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
        self.load_cache()

    def clear_all(self):
        if QMessageBox.question(
            self, "전체 초기화", "캐시를 모두 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No
        ) != QMessageBox.Yes:
            return
        if os.path.exists(CACHE_FILE):
            os.remove(CACHE_FILE)
        self.load_cache()


class VirusScannerTab(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.setWindowTitle("💻 IP 분석 자동화 도구")
        self.setGeometry(100, 100, 800, 600)
        self.setFont(QFont("Segoe UI", 10))

        self.file_paths = {"AIPS": None, "HIPS": None, "웹방화벽": None}
        self.progress_bars = {}
        self.status_labels = {}
        self.mode_checkboxes = {}
        self.setStyleSheet("""
    QWidget {
        background-color: #f4f6f9;
        font-size: 13px;
    }
    QLabel {
        font-size: 13px;
    }
    QPushButton {
    background-color: #1976d2;
    color: white;
    border-radius: 6px;
    padding: 10px 20px;
    font-size: 13px;
    font-weight: bold;
}
QPushButton:hover {
    background-color: #1565c0;
}
QLabel {
    font-size: 13px;
}
QTextEdit {
    background-color: #f5f5f5;
    border: 1px solid #ccc;
    padding: 8px;
    font-size: 13px;
}

    QProgressBar {
        height: 20px;
        border: 1px solid #aaa;
        border-radius: 10px;
        text-align: center;
        font-weight: bold;
    }
    QProgressBar::chunk {
        background-color: #4caf50;
        border-radius: 10px;
    }
    QGroupBox {
        border: 1px solid #ccc;
        border-radius: 6px;
        margin-top: 10px;
        padding: 10px;
        background-color: #ffffff;
    }
    QGroupBox:title {
        subcontrol-origin: margin;
        left: 10px;
        padding: 0 3px;
        font-size: 14px;
        font-weight: bold;
    }
   QCheckBox {
    spacing: 12px;
    font-size: 15px;
    padding: 6px;
    color: #333;
    font-weight: 500;
}

QCheckBox::indicator {
    width: 24px;
    height: 24px;
    border-radius: 4px;
    border: 2px solid #b0bec5;
    background-color: #ffffff;
    margin-right: 6px;
}

QCheckBox::indicator:checked {
    background-color: #c8e6c9;
    border: 2px solid #2e7d32;
}


                           
    QPushButton#runButton {
        background-color: #43a047;  
    }
    QPushButton#runButton:hover {
        background-color: #39b091;
    }
    
""")

        layout = QVBoxLayout()
        layout.setContentsMargins(30, 20, 30, 20)
        layout.setSpacing(10)
        self.setLayout(layout)

        title = QLabel("🔎 IP 위협 정보 분석기 (모드별 개별 분석)")
        title.setFont(QFont("Segoe UI", 18, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        for mode in ["AIPS", "HIPS", "웹방화벽"]:
            layout.addSpacing(10)

            # 🔹 모드 선택 체크박스
            cb = QCheckBox(f"{mode} 분석 실행")
            cb.setChecked(False)
            self.mode_checkboxes[mode] = cb
            layout.addWidget(cb)

            # 🔹 파일 선택 버튼 + 파일 라벨
            file_btn = QPushButton(f"{mode} CSV 선택")
            file_btn.clicked.connect(lambda _, m=mode: self.select_file(m))
            layout.addWidget(file_btn)

            file_label = QLabel("📂 현재 파일 없음")
            file_label.setObjectName(f"file_label_{mode}")
            layout.addWidget(file_label)

            # 🔹 진행률 바
            progress = QProgressBar()
            progress.setValue(0)
            self.progress_bars[mode] = progress
            layout.addWidget(progress)

            # 🔹 상태 메시지
            status = QLabel("")
            status.setObjectName(f"status_label_{mode}")
            self.status_labels[mode] = status
            layout.addWidget(status)


        # 분석 실행 버튼
        self.btn_run = QPushButton("🚀 선택된 분석 실행")
        self.btn_run.setObjectName("runButton")
        self.btn_run.setFixedHeight(40)
        self.btn_run.clicked.connect(self.run_analysis)
        layout.addWidget(self.btn_run)


    def select_file(self, mode):
        fname, _ = QFileDialog.getOpenFileName(self, f"{mode} CSV 파일 선택", "", "CSV Files (*.csv)")
        if fname:
            self.file_paths[mode] = fname
            label = self.findChild(QLabel, f"file_label_{mode}")
            if label:
                label.setText(f"📂 선택된 파일: {fname}")



    def load_file(self):
        fname, _ = QFileDialog.getOpenFileName(self, "CSV 파일 선택", "", "CSV Files (*.csv)")
        if fname:
            self.file_path = fname
            self.file_label.setText(f"선택된 파일: {fname}")

    def run_analysis(self):
        # 실행 대상 모드: 체크박스 + 파일 둘 다 만족해야 함
        selected_modes = [
            mode for mode in self.file_paths
            if self.mode_checkboxes[mode].isChecked() and self.file_paths[mode]
        ]

        if not selected_modes:
            QMessageBox.warning(self, "오류", "하나 이상의 모드를 선택하고 파일도 지정해주세요.")
            return

        self.threads = []

        for mode in selected_modes:
            path = self.file_paths[mode]

            self.status_labels[mode].setText("🔄 분석 준비 중...")
            self.progress_bars[mode].setValue(0)

            thread = AnalysisThread(path, mode)
            thread.progress.connect(lambda p, ip, m=mode: self.update_progress(p, ip, m))
            thread.finished.connect(lambda msg, file, m=mode: self.on_analysis_done(msg, file, m))
            thread.error.connect(lambda err, m=mode: self.on_analysis_error(err, m))

            self.threads.append(thread)
            thread.start()


    def update_progress(self, percent, ip, mode):
        bar = self.progress_bars.get(mode)
        label = self.status_labels.get(mode)
        thread = next((t for t in self.threads if t.mode == mode), None)

        if bar:
            bar.setValue(percent)
        if label and thread:
            label.setText(f"🔍 [{mode}] {ip} 조회 중 ({thread.current_index + 1}/{thread.total_count}, {percent}%)")


    def on_analysis_done(self, stats_msg, output_file, mode):
        self.status_labels[mode].setText("✅ 완료")
        self.main_window.result_files.append({"path":output_file,  "source": mode})
        QMessageBox.information(self, f"{mode} 분석 완료", f"{stats_msg}\n📁 저장 파일: {output_file}")

        line = f"{datetime.now():%Y-%m-%d %H:%M:%S} [{mode}] {stats_msg}"
        self.main_window.history_tab.append_entry(line)

        log_line = f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} [{mode}] {stats_msg}\n"
        with open("analysis_history.log", "a", encoding="utf-8") as lf:
            lf.write(log_line)

    def on_analysis_error(self, error_msg, mode):
        self.status_labels[mode].setText("❌ 오류 발생")
        QMessageBox.critical(self, f"{mode} 오류", error_msg)



class JsonToExcelTab(QWidget):
    def __init__(self, main_window):
        super().__init__()
        
        self.file_path = None
        self.main_window = main_window

        layout = QVBoxLayout()
        layout.setContentsMargins(40, 30, 40, 30)
        layout.setSpacing(15)
        self.setLayout(layout)

        self.setStyleSheet("""
            QPushButton {
                background-color: #1976d2;
                color: white;
                border-radius: 6px;
                padding: 10px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1565c0;
            }
            QLabel {
                font-size: 20px;
            }
        """)

        self.label = QLabel("📂 JSON 파일을 선택하세요")
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label)

        self.btn_select = QPushButton("📁 JSON 파일 선택")
        self.btn_select.clicked.connect(self.select_file)
        layout.addWidget(self.btn_select)

        self.btn_convert = QPushButton("🚀 변환 실행")
        self.btn_convert.setEnabled(False)
        self.btn_convert.clicked.connect(self.convert)
        layout.addWidget(self.btn_convert)

    def select_file(self):
        fname, _ = QFileDialog.getOpenFileName(self, "JSON 파일 선택", "", "JSON Files (*.json)")
        if fname:
            self.file_path = fname
            self.label.setText(f"📄 선택된 파일:\n{fname}")
            self.btn_convert.setEnabled(True)

    def convert(self):
        try:
            with open(self.file_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)

            filtered_data = [
                {
                    "위협 IP": item.get("threat_info"),
                    "정책 구분": item.get("policy_div"),
                    "위협 사유": item.get("reason_div").strip() if item.get("reason_div") else None
                }
                for item in json_data.get("datas", [])
            ]

            df = pd.DataFrame(filtered_data)
            output_file = self.file_path.replace('.json', '_output.xlsx')
            df.to_excel(output_file, index=False)

            wb = load_workbook(output_file)
            ws = wb.active
            for col_num, column_cells in enumerate(ws.columns, 1):
                max_len = max((len(str(cell.value)) if cell.value else 0) for cell in column_cells)
                for cell in column_cells:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.column_dimensions[get_column_letter(col_num)].width = max_len + 4
            for cell in ws[1]:
                cell.font = ExcelFont(bold=True)
            wb.save(output_file)

            self.main_window.result_files.append(output_file)  # ✅ 수정된 부분
            QMessageBox.information(self, "✅ 완료", f"Excel로 변환됨:\n{output_file}")
        except Exception as e:
            QMessageBox.critical(self, "오류", str(e))


class CLIGeneratorTab(QWidget):
    def __init__(self, main_window ):
        super().__init__()
        self.main_window = main_window 
        self.input_file = ""
        self.output_file = "fortigate_block_ips.txt"
        self.cache_file = "group_cache.json"
        self.group_capacity = 300

        self.setFont(QFont("맑은 고딕", 10))
        self.setStyleSheet("""
            QPushButton {
                background-color: #1976d2;
                color: white;
                border-radius: 6px;
                padding: 10px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1565c0;
            }
            QLabel {
                margin-bottom: 5px;
                font-size: 20px;
                margin-left: 160px;
            }
            QTextEdit {
                background-color: #f5f5f5;
                border: 1px solid #ccc;
                padding: 6px;
            }
        """)

        layout = QVBoxLayout()
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(10)
        self.setLayout(layout)

        self.label = QLabel("📄 엑셀 파일을 선택하세요")
        layout.addWidget(self.label)

        self.select_button = QPushButton("엑셀 파일 선택")
        self.select_button.clicked.connect(self.select_file)
        layout.addWidget(self.select_button)

        self.generate_button = QPushButton("CLI 생성 실행")
        self.generate_button.clicked.connect(self.generate_cli)
        layout.addWidget(self.generate_button)

        self.reset_button = QPushButton("캐시 초기화")
        self.reset_button.clicked.connect(self.reset_cache)
        layout.addWidget(self.reset_button)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setFixedHeight(180)
        layout.addWidget(self.log_box)

    def log(self, message):
        self.log_box.append(message)
        self.log_box.moveCursor(QTextCursor.End)

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "엑셀 파일 선택", "", "Excel Files (*.xlsx)")
        if file_path:
            self.input_file = file_path
            self.label.setText(f"📄 선택된 파일:\n{file_path}")

    def reset_cache(self):
        if os.path.exists(self.cache_file):
            os.remove(self.cache_file)
            self.log("✅ 그룹 캐시 초기화 완료")
        else:
            self.log("ℹ️ 캐시 파일 없음")

    def generate_cli(self):
        if not self.input_file:
            QMessageBox.warning(self, "경고", "엑셀 파일을 먼저 선택하세요.")
            return

        try:
            df = pd.read_excel(self.input_file)
            ip_list = df['위협 IP'].dropna().unique().tolist()
        except Exception as e:
            self.log(f"❌ 엑셀 파일 오류: {e}")
            return

        file_base = os.path.splitext(os.path.basename(self.input_file))[0]
        base_output_name = f"{file_base}_CLI"
        file_count = 1

        def save_cli(lines):
            nonlocal file_count
            fname = f"{base_output_name}_{file_count}.txt"
            with open(fname, 'w', encoding='utf-8') as f:
                for line in lines:
                    f.write(line + '\n')
            self.log(f"✅ CLI 저장됨: {fname}")
            file_count += 1

        if os.path.exists(self.cache_file):
            with open(self.cache_file, 'r') as f:
                cache = json.load(f)
            group_name = cache.get("group_name", "")
            current_count = cache.get("count", 0)
            ip_index = cache.get("ip_index", 0)
        else:
            group_name = ""
            current_count = 0
            ip_index = 0

        if not group_name:
            group_name_input, ok = QInputDialog.getText(self, "그룹명 입력", "FortiGate 그룹명을 입력하세요:")
            if not ok or not group_name_input.strip():
                self.log("❌ 그룹명이 입력되지 않아 중단됨")
                return
            group_name = group_name_input.strip()

        cli_lines = []
        ip_name_list = []
        cli_lines.append("config firewall address")

        for ip in ip_list:
            if current_count >= self.group_capacity:
                self.log(f"⚠ 그룹 '{group_name}'에 300개 도달")
                cli_lines.append("end")
                cli_lines.append("")
                cli_lines.append("config firewall addrgrp")
                cli_lines.append(f'edit "{group_name}"')
                cli_lines.append("append member " + ' '.join(f'"{ip}"' for ip in ip_name_list))
                cli_lines.append("next")
                cli_lines.append("end\n")

                save_cli(cli_lines)

                group_name_input, ok = QInputDialog.getText(self, "새 그룹명 입력", "새 FortiGate 그룹명을 입력하세요:")
                if not ok or not group_name_input.strip():
                    self.log("❌ 그룹명이 입력되지 않아 중단됨")
                    break
                group_name = group_name_input.strip()
                current_count = 0
                ip_name_list = []
                cli_lines = ["config firewall address"]

            ip_index += 1
            ip_name = f"C-TAS-ip-{ip_index}"
            ip_name_list.append(ip_name)
            cli_lines.append(f'edit "{ip_name}"')
            cli_lines.append(f'set subnet {ip} 255.255.255.255')
            cli_lines.append("next")
            current_count += 1

        if ip_name_list:
            cli_lines.append("end\n")
            cli_lines.append("config firewall addrgrp")
            cli_lines.append(f'edit "{group_name}"')
            cli_lines.append("append member " + ' '.join(f'"{ip}"' for ip in ip_name_list))
            cli_lines.append("next")
            cli_lines.append("end\n")

            save_cli(cli_lines)

        with open(self.cache_file, 'w') as f:
            json.dump({
                "group_name": group_name,
                "count": current_count,
                "ip_index": ip_index
            }, f, indent=2)

        self.log(f"📌 마지막 그룹: {group_name}, IP: {current_count}/300")
        QMessageBox.information(self, "완료", "CLI 파일 생성 완료")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("💻 보안 자동화 도구 통합판")
        self.setGeometry(200, 200, 640, 500)
        self.result_files = []  # 📁 생성된 결과 파일들 기록
        


        self.tabs = QTabWidget()

        self.tabs.setStyleSheet("""
    QTabBar::tab {
        font-size: 15px;
        padding: 8px 20px;
        height: 40px;
        background: #e0e0e0;
        border-top-left-radius: 8px;
        border-top-right-radius: 8px;
        margin-right: 3.3px;
        width:81.2px
    }
    QTabBar::tab:selected {
        background: #1976d2;
        color: white;
        font-weight: bold;
    }
    QTabWidget::pane {
        border-top: 2px solid #1976d2;
        top: -1px;
    }
""")
        
        self.json_tab = JsonToExcelTab(self)
        self.virus_tab = VirusScannerTab(self)
        self.tabs.setFont(QFont("맑은 고딕", 11))
        self.history_tab = HistoryTab(self)
        self.cache_tab = CacheTab()
        self.tabs.addTab(self.json_tab, "📄 C-TAS")
        self.tabs.addTab(CLIGeneratorTab(self), "🛡️ CLI 생성")
        self.tabs.addTab(self.virus_tab, "🔬 IP 분석")
        self.tabs.addTab(self.history_tab, "📜 로그")
        self.tabs.addTab(self.cache_tab, "🗄️ 캐시")


        

        self.setCentralWidget(self.tabs)


        # 하단 버튼 영역
        main_layout = QVBoxLayout()
        main_layout.addWidget(self.tabs)

        btn_export_all = QPushButton("📤 통합 위협 IP 저장")
        btn_export_all.clicked.connect(self.export_all_ips)
        btn_export_all.setStyleSheet("""
    QPushButton {
        background-color: #FF9800;   /* 녹색 계열 */
        color: white;
        padding: 8px 16px;
        border: none;
        border-radius: 6px;
        font-weight: bold;
    }
    QPushButton:hover {
        background-color: #FB8C00;
    }
    QPushButton:pressed {
        background-color: #EF6C00;
    }
""")
        main_layout.addWidget(btn_export_all)

        wrapper = QWidget()
        wrapper.setLayout(main_layout)
        self.setCentralWidget(wrapper)

    def detect_source(self, file_path, df):
        lower = file_path.lower()
        if "위협 IP" in df.columns and "위협 사유" in df.columns:
            return "C-TAS"
        elif "IP" in df.columns and "Attack_Type" in df.columns:
            if "aips" in lower:
                return "AIPS"
            elif "hips" in lower:
                return "HIPS"
            else:
                return "AIPS/HIPS"
        elif "출발지 주소" in df.columns and "룰" in df.columns:
            return "웹방화벽"
        else:
            return "알수없음"

    def export_all_ips(self):
        new_data = []

        # 1) 각 결과 파일에서 데이터 수집할 때 Malicious 값도 함께 담기
        for entry in self.result_files:
            file_path = entry.get("path") if isinstance(entry, dict) else entry
            forced_source = entry.get("source") if isinstance(entry, dict) else None

            try:
                df = pd.read_excel(file_path)
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                source = forced_source or self.detect_source(file_path, df)

                for _, row in df.iterrows():
                    malicious = row.get("Malicious", 0)
                    if "위협 IP" in df.columns and "위협 사유" in df.columns:
                        ip = row["위협 IP"]
                        ttype = row.get("위협 사유", "")
                    elif "IP" in df.columns and "Attack_Type" in df.columns:
                        ip = row["IP"]
                        ttype = row.get("Attack_Type", "")
                    elif "출발지 주소" in df.columns and "룰" in df.columns:
                        ip = row["출발지 주소"]
                        ttype = row.get("룰", "")
                    else:
                        continue

                    new_data.append({
                        "출처": source,
                        "위협 IP": ip,
                        "위협 유형": ttype,
                        "분석 일시": now,
                        "Malicious": malicious
                    })

            except Exception as e:
                QMessageBox.warning(self, "⚠️ 읽기 실패", f"{file_path} 에서 오류: {e}")

        if not new_data:
            QMessageBox.information(self, "📭 없음", "수집된 IP가 없습니다.")
            return

        # 2) DataFrame 생성
        df_new = pd.DataFrame(new_data)
        output_file = os.path.join(os.getcwd(), "all_threat_ips.xlsx")

        # 3) 기존 파일이 있으면 합치기
        if os.path.exists(output_file):
            try:
                df_existing = pd.read_excel(output_file)
                df_all = pd.concat([df_existing, df_new], ignore_index=True)
            except Exception as e:
                QMessageBox.critical(self, "❌ 파일 읽기 오류", f"기존 통합 파일을 열 수 없습니다:\n{e}")
                return
        else:
            df_all = df_new

        # 4) 출처 문자열화
        df_all["출처"] = df_all["출처"].fillna("").astype(str)

        # 5) 그룹핑하면서 집계: 출처는 merge, 분석 일시는 최신, Malicious는 최대
        def merge_sources(series):
            seen = []
            for item in series:
                for src in re.split(r'\s*,\s*', str(item)):
                    s = src.strip()
                    if s and s not in seen:
                        seen.append(s)
            return ", ".join(seen)

        df_all = (
            df_all
            .groupby(["위협 IP", "위협 유형"], as_index=False)
            .agg({
                "출처": merge_sources,
                "분석 일시": "max",
                "Malicious": "max"
            })
        )

        # 6) 차단됨 컬럼 추가:
        #    - C-TAS 출처면 무조건 'V'
        #    - 그 외에는 Malicious ≥ 1 → 'V'
        def is_blocked(row):
            if "C-TAS" in row["출처"]:
                return "V"
            return "V" if row["Malicious"] >= 1 else ""

        df_all["차단됨"] = df_all.apply(is_blocked, axis=1)

        # 7) 컬럼 순서 설정 및 Malicious 컬럼 제거
        cols = ["출처", "위협 IP", "위협 유형", "분석 일시", "차단됨"]
        df_all = df_all[[c for c in cols if c in df_all.columns]]

        # 8) 엑셀로 저장 및 열 너비 자동 조정
        try:
            df_all.to_excel(output_file, index=False)
            QMessageBox.information(self, "✅ 저장 완료", f"통합 IP 목록 저장됨:\n{output_file}")
            wb = load_workbook(output_file)
            ws = wb.active
            for column_cells in ws.columns:
                max_len = max((len(str(cell.value)) if cell.value else 0) for cell in column_cells)
                col_letter = get_column_letter(column_cells[0].column)
                ws.column_dimensions[col_letter].width = max_len + 4
            wb.save(output_file)
        except Exception as e:
            QMessageBox.critical(self, "❌ 저장 실패", str(e))


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
